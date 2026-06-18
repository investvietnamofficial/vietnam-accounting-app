"""
Reports API - VAT declarations, annex lists, and provisional CIT calculations.

The backend now computes declaration fields using the official 01/GTGT formulas
instead of the earlier MVP aggregates.
"""

from __future__ import annotations

from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.security import get_current_user
from app.models import (
    Company,
    Invoice,
    JournalEntry,
    JournalEntryStatus,
    VATRate,
)
from app.services.tax.vn_tax_engine import (
    CIT_STANDARD_RATE,
    InvoiceIssue,
    calculate_quarterly_cit_provision,
    detect_duplicate_invoices,
    detect_low_confidence,
    detect_missing_mst,
    detect_vat_mismatch,
    get_cit_quarter_payment_deadline,
    get_vat_declaration_deadline,
    normalize_tax_code,
    validate_mst,
)

router = APIRouter()

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _period_range(year: int, period: int, period_type: str) -> tuple[datetime, datetime]:
    if period_type == "monthly":
        if period < 1 or period > 12:
            raise HTTPException(status_code=400, detail="Monthly period must be 1-12")
        start = date(year, period, 1)
        end = date(year + (period == 12), 1 if period == 12 else period + 1, 1)
    else:
        if period < 1 or period > 4:
            raise HTTPException(status_code=400, detail="Quarterly period must be 1-4")
        start_month = (period - 1) * 3 + 1
        start = date(year, start_month, 1)
        end = date(year + (period == 4), 1 if period == 4 else start_month + 3, 1)
    return datetime.combine(start, time.min), datetime.combine(end, time.min)


def _quarter_end_date(year: int, quarter: int) -> datetime:
    _, end = _period_range(year, quarter, "quarterly")
    return end


async def _company(db: AsyncSession, company_id: str) -> Company | None:
    result = await db.execute(select(Company).where(Company.id == company_id))
    return result.scalar_one_or_none()


async def _period_invoices(
    db: AsyncSession,
    company_id: str,
    start_date: datetime,
    end_date: datetime,
) -> list[Invoice]:
    result = await db.execute(
        select(Invoice)
        .where(Invoice.company_id == company_id)
        .where(Invoice.invoice_date >= start_date)
        .where(Invoice.invoice_date < end_date)
        .order_by(Invoice.invoice_date.asc(), Invoice.invoice_number.asc())
    )
    return result.scalars().all()


async def _ytd_posted_journal_entries(
    db: AsyncSession,
    company_id: str,
    year: int,
    end_date: datetime,
) -> list[JournalEntry]:
    start_date = datetime.combine(date(year, 1, 1), time.min)
    result = await db.execute(
        select(JournalEntry)
        .options(selectinload(JournalEntry.lines))
        .where(JournalEntry.company_id == company_id)
        .where(JournalEntry.status == JournalEntryStatus.POSTED)
        .where(JournalEntry.entry_date >= start_date)
        .where(JournalEntry.entry_date < end_date)
        .order_by(JournalEntry.entry_date.asc())
    )
    return result.scalars().all()


def _invoice_direction(inv: Invoice, company: Company) -> str:
    company_tax_code = normalize_tax_code(company.tax_code)
    seller_tax_code = normalize_tax_code(inv.seller_tax_code)
    if company_tax_code and seller_tax_code == company_tax_code:
        return "sale"
    return "purchase"


def _period_label(period_type: str, period: int, year: int) -> str:
    if period_type == "monthly":
        return f"Thang {period:02d}/{year}"
    return f"Quy {period}/{year}"


def _build_annexes(invoices: list[Invoice], company: Company) -> tuple[dict, dict]:
    purchase_items: list[dict] = []
    sales_items: list[dict] = []
    purchase_totals = {"count": 0, "taxable_value": 0, "vat_amount": 0, "total_amount": 0}
    sales_totals = {"count": 0, "taxable_value": 0, "vat_amount": 0, "total_amount": 0}

    for index, invoice in enumerate(invoices, start=1):
        row = _invoice_report_row(index, invoice, company)
        if row["direction"] == "purchase":
            purchase_items.append(row)
            purchase_totals["count"] += 1
            purchase_totals["taxable_value"] += row["subtotal_amount"]
            purchase_totals["vat_amount"] += row["vat_amount"]
            purchase_totals["total_amount"] += row["total_amount"]
        else:
            sales_items.append(row)
            sales_totals["count"] += 1
            sales_totals["taxable_value"] += row["subtotal_amount"]
            sales_totals["vat_amount"] += row["vat_amount"]
            sales_totals["total_amount"] += row["total_amount"]

    return (
        {
            "code": "01-1/GTGT",
            "title": "Phu luc bang ke hoa don mua vao",
            "items": purchase_items,
            "totals": purchase_totals,
        },
        {
            "code": "01-2/GTGT",
            "title": "Phu luc bang ke hoa don ban ra",
            "items": sales_items,
            "totals": sales_totals,
        },
    )


def _invoice_report_row(index: int, inv: Invoice, company: Company) -> dict:
    counterparty_name = inv.seller_name if _invoice_direction(inv, company) == "purchase" else inv.buyer_name
    counterparty_tax_code = inv.seller_tax_code if _invoice_direction(inv, company) == "purchase" else inv.buyer_tax_code
    return {
        "stt": index,
        "id": inv.id,
        "direction": _invoice_direction(inv, company),
        "invoice_date": inv.invoice_date.date().isoformat() if inv.invoice_date else None,
        "invoice_series": inv.invoice_series,
        "invoice_number": inv.invoice_number,
        "counterparty_name": counterparty_name,
        "counterparty_tax_code": counterparty_tax_code,
        "seller_name": inv.seller_name,
        "seller_tax_code": inv.seller_tax_code,
        "buyer_name": inv.buyer_name,
        "buyer_tax_code": inv.buyer_tax_code,
        "subtotal_amount": int(inv.subtotal_amount or 0),
        "vat_rate": float(str(inv.vat_rate.value) or "10") / 100,
        "vat_amount": int(inv.vat_amount or 0),
        "total_amount": int(inv.total_amount or 0),
        "einvoice_verified": inv.einvoice_verified,
        "confidence": float(inv.extraction_confidence) if inv.extraction_confidence is not None else None,
    }


def _aggregate_cit_bases(journal_entries: list[JournalEntry]) -> tuple[int, int, int, int]:
    revenue = 0
    deductible_expenses = 0
    other_income = 0
    other_expenses = 0

    for entry in journal_entries:
        for line in entry.lines:
            amount = int(line.amount or 0)
            if line.credit_account_code:
                credit_code = str(line.credit_account_code)
                if credit_code.startswith("511") or credit_code.startswith("515"):
                    revenue += amount
                elif credit_code.startswith("711"):
                    other_income += amount
            if line.debit_account_code:
                debit_code = str(line.debit_account_code)
                if debit_code.startswith(("632", "635", "641", "642")) or debit_code.startswith("6"):
                    deductible_expenses += amount
                elif debit_code.startswith("811"):
                    other_expenses += amount

    return revenue, deductible_expenses, other_income, other_expenses


def _invoice_list_payload(
    invoices: list[Invoice],
    company: Company,
    period_type: str,
    period: int,
    year: int,
    direction_filter: str | None = None,
) -> dict:
    """Build invoice list for sales/purchase report endpoints."""
    if direction_filter:
        invoices = [inv for inv in invoices if _invoice_direction(inv, company) == direction_filter]

    items = []
    for idx, inv in enumerate(invoices, start=1):
        items.append({
            "id": inv.id,
            "invoice_number": inv.invoice_number,
            "invoice_date": inv.invoice_date.date().isoformat() if inv.invoice_date else None,
            "seller_name": inv.seller_name,
            "seller_tax_code": inv.seller_tax_code,
            "buyer_name": inv.buyer_name,
            "buyer_tax_code": inv.buyer_tax_code,
            "subtotal": int(inv.subtotal_amount or 0),
            "vat_rate": float(str(inv.vat_rate.value) or "10") / 100,
            "vat_amount": int(inv.vat_amount or 0),
            "total": int(inv.total_amount or 0),
            "confidence": float(inv.extraction_confidence) if inv.extraction_confidence is not None else None,
        })

    return {
        "items": items,
        "total": len(items),
        "period_label": _period_label(period_type, period, year),
    }


def _build_vat_workbook(summary: dict, company: Company) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    declaration_sheet = wb.active
    declaration_sheet.title = "01-GTGT"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    number_format = "#,##0"

    declaration_sheet.merge_cells("A1:E1")
    declaration_sheet["A1"] = "TO KHAI THUE GIA TRI GIA TANG - MAU 01/GTGT"
    declaration_sheet["A1"].font = title_font
    declaration_sheet["A1"].alignment = center
    declaration_sheet.append(["Ten nguoi nop thue", company.name, "", "MST", company.tax_code])
    declaration_sheet.append(["Ky khai thue", _period_label(summary["period_type"], summary["period"], summary["year"]), "", "Han nop", summary.get("filing_deadline", "")])
    declaration_sheet.append([])

    declaration_sheet.append(["STT", "Chi tieu", "Ma", "Gia tri HHDV chua thue", "Thue GTGT"])
    for cell in declaration_sheet[5]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    filing_fields = summary["filing_fields"]
    declaration_rows = [
        ("A", 'Khong phat sinh hoat dong mua, ban trong ky', "[21]", "X" if filing_fields["21"] else "", ""),
        ("B", "Thue GTGT con duoc khau tru ky truoc chuyen sang", "[22]", "", filing_fields["22"]),
        ("C1", "Gia tri HHDV mua vao trong ky", "[23]", filing_fields["23"], filing_fields["24"]),
        ("C1a", "Trong do: hang hoa, dich vu nhap khau", "[23a]", filing_fields["23a"], filing_fields["24a"]),
        ("C2", "Thue GTGT mua vao duoc khau tru ky nay", "[25]", "", filing_fields["25"]),
        ("D1", "Ban ra khong chiu thue GTGT", "[26]", filing_fields["26"], ""),
        ("D2", "Ban ra chiu thue GTGT", "[27]", filing_fields["27"], filing_fields["28"]),
        ("D2a", "Ban ra thue suat 0%", "[29]", filing_fields["29"], ""),
        ("D2b", "Ban ra thue suat 5%", "[30]", filing_fields["30"], filing_fields["31"]),
        ("D2c", "Ban ra thue suat tieu chuan (10% va 8% theo HTKK)", "[32]", filing_fields["32"], filing_fields["33"]),
        ("D2d", "Ban ra khong phai ke khai, tinh nop thue", "[32a]", filing_fields["32a"], ""),
        ("D3", "Tong doanh thu va thue GTGT ban ra", "[34]", filing_fields["34"], filing_fields["35"]),
        ("E", "Thue GTGT phat sinh trong ky", "[36]", "", filing_fields["36"]),
        ("F1", "Dieu chinh giam thue GTGT duoc khau tru ky truoc", "[37]", "", filing_fields["37"]),
        ("F2", "Dieu chinh tang thue GTGT duoc khau tru ky truoc", "[38]", "", filing_fields["38"]),
        ("G", "Thue GTGT nhan ban giao duoc khau tru trong ky", "[39a]", "", filing_fields["39a"]),
        ("H1", "Thue GTGT phai nop cua hoat dong SXKD trong ky", "[40a]", "", filing_fields["40a"]),
        ("H2", "Thue GTGT du an dau tu duoc bu tru", "[40b]", "", filing_fields["40b"]),
        ("H3", "Thue GTGT con phai nop trong ky", "[40]", "", filing_fields["40"]),
        ("H4", "Thue GTGT chua khau tru het ky nay", "[41]", "", filing_fields["41"]),
        ("H4.1", "Thue GTGT de nghi hoan", "[42]", "", filing_fields["42"]),
        ("H4.2", "Thue GTGT con duoc khau tru chuyen ky sau", "[43]", "", filing_fields["43"]),
    ]

    for row in declaration_rows:
        declaration_sheet.append(list(row))
        current_row = declaration_sheet.max_row
        for cell in declaration_sheet[current_row]:
            cell.border = thin_border
            if cell.column in (4, 5):
                cell.alignment = right
                if isinstance(cell.value, int):
                    cell.number_format = number_format

    declaration_sheet.column_dimensions["A"].width = 10
    declaration_sheet.column_dimensions["B"].width = 54
    declaration_sheet.column_dimensions["C"].width = 10
    declaration_sheet.column_dimensions["D"].width = 20
    declaration_sheet.column_dimensions["E"].width = 20

    _append_annex_sheet(wb, "01-1-GTGT", summary["purchase_annex"])
    _append_annex_sheet(wb, "01-2-GTGT", summary["sales_annex"])

    # VAT Summary sheet
    summary_ws = wb.create_sheet("VAT Summary")
    summary_ws.append(["VAT Summary Report"])
    summary_ws["A1"].font = title_font
    summary_ws.append(["Company Name", company.name])
    summary_ws.append(["Tax Code", company.tax_code])
    summary_ws.append(["Period", _period_label(summary["period_type"], summary["period"], summary["year"])])
    summary_ws.append(["Generated At", summary.get("generated_at", datetime.utcnow().isoformat())])
    summary_ws.append([])

    summary_ws.append(["", "Input VAT (VND)", "Output VAT (VND)"])
    summary_ws.append(["Total", summary["input_vat_total"], summary["output_vat_total"]])
    summary_ws.append(["Net VAT", summary["net_vat"], ""])
    summary_ws.append([])

    summary_ws.append(["VAT Rate", "Input Amount (VND)", "Output Amount (VND)", "Input VAT (VND)", "Output VAT (VND)"])
    by_rate = {str(r["rate"]): r for r in summary.get("by_rate", [])}
    for rate, label in [("0", "0%"), ("5", "5%"), ("8", "8%"), ("10", "10%")]:
        row_data = by_rate.get(rate, {})
        summary_ws.append([
            label,
            row_data.get("input_amount", 0),
            row_data.get("output_amount", 0),
            row_data.get("input_vat", 0),
            row_data.get("output_vat", 0),
        ])

    summary_ws.append([])
    summary_ws.append(["Invoice Count", summary.get("invoice_count", 0)])
    summary_ws.append(["Filing Deadline", summary.get("filing_deadline", "")])

    if summary.get("warnings"):
        summary_ws.append([])
        summary_ws.append(["Warnings"])
        for w in summary["warnings"]:
            summary_ws.append([w["type"], w["message"]])

    for col, width in [("A", 20), ("B", 22), ("C", 22), ("D", 22), ("E", 22)]:
        summary_ws.column_dimensions[col].width = width

    # Validation issues sheet
    if summary["validation_issues"]:
        issues_sheet = wb.create_sheet("Validation")
        issues_sheet.append(["Validation issues"])
        issues_sheet["A1"].font = bold
        for issue in summary["validation_issues"]:
            issues_sheet.append([issue])
        issues_sheet.column_dimensions["A"].width = 120

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _append_annex_sheet(wb, title: str, annex: dict) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    ws = wb.create_sheet(title)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    number_format = "#,##0"

    ws.append([annex["title"]])
    ws["A1"].font = Font(bold=True, size=13)
    headers = [
        "STT",
        "Ngay hoa don",
        "Ky hieu",
        "So hoa don",
        "Ten doi tac",
        "MST doi tac",
        "Gia tri chua thue",
        "Thue suat",
        "Tien thue GTGT",
        "Tong thanh toan",
    ]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = bold
        cell.alignment = center
        cell.fill = header_fill
        cell.border = thin_border

    for item in annex["items"]:
        ws.append(
            [
                item["stt"],
                item["invoice_date"] or "",
                item["invoice_series"] or "",
                item["invoice_number"] or "",
                item["counterparty_name"] or "",
                item["counterparty_tax_code"] or "",
                item["subtotal_amount"],
                item["vat_rate"],
                item["vat_amount"],
                item["total_amount"],
            ]
        )
        row_index = ws.max_row
        for cell in ws[row_index]:
            cell.border = thin_border
            if cell.column in (7, 9, 10):
                cell.number_format = number_format
                cell.alignment = right

    ws.append(
        [
            "",
            "",
            "",
            "",
            "Tong cong",
            "",
            annex["totals"]["taxable_value"],
            "",
            annex["totals"]["vat_amount"],
            annex["totals"]["total_amount"],
        ]
    )
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = bold
        cell.border = thin_border
        if cell.column in (7, 9, 10):
            cell.number_format = number_format
            cell.alignment = right

    widths = {"A": 8, "B": 14, "C": 14, "D": 14, "E": 34, "F": 18, "G": 18, "H": 12, "I": 18, "J": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _build_invoice_list_workbook(items: list[dict], title: str, company_name: str) -> BytesIO:
    """Build XLSX for sales or purchase invoice list."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    number_format = "#,##0"
    date_format = "DD/MM/YYYY"

    ws.append(["Invoice List - " + title, "", "", "", "", "", "", "", "", "", ""])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Company", company_name])
    ws.append([])

    headers = [
        "Invoice #",
        "Date",
        "Seller",
        "Seller MST",
        "Buyer",
        "Buyer MST",
        "Subtotal (VND)",
        "VAT Rate",
        "VAT Amount (VND)",
        "Total (VND)",
        "Confidence",
    ]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = bold
        cell.alignment = center
        cell.fill = header_fill
        cell.border = thin_border

    for item in items:
        inv_date = item.get("invoice_date") or ""
        # Format date as DD/MM/YYYY
        if inv_date and isinstance(inv_date, str) and len(inv_date) == 10:
            parts = inv_date.split("-")
            if len(parts) == 3:
                inv_date = f"{parts[2]}/{parts[1]}/{parts[0]}"

        ws.append([
            item.get("invoice_number") or "",
            inv_date,
            item.get("seller_name") or "",
            item.get("seller_tax_code") or "",
            item.get("buyer_name") or "",
            item.get("buyer_tax_code") or "",
            item.get("subtotal") or 0,
            item.get("vat_rate") or 0,
            item.get("vat_amount") or 0,
            item.get("total") or 0,
            f"{item.get('confidence', 0) * 100:.0f}%" if item.get("confidence") is not None else "",
        ])
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.border = thin_border
            if cell.column in (7, 9, 10):
                cell.number_format = number_format
                cell.alignment = right
            elif cell.column == 8:
                cell.number_format = "0%"
                cell.alignment = right

    # Totals row
    total_subtotal = sum(i.get("subtotal") or 0 for i in items)
    total_vat = sum(i.get("vat_amount") or 0 for i in items)
    total_amount = sum(i.get("total") or 0 for i in items)
    ws.append(["", "", "", "", "TOTAL", "", total_subtotal, "", total_vat, total_amount, ""])
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = bold
        cell.border = thin_border
        if cell.column in (7, 9, 10):
            cell.number_format = number_format
            cell.alignment = right

    widths = {"A": 16, "B": 14, "C": 32, "D": 18, "E": 32, "F": 18, "G": 18, "H": 12, "I": 18, "J": 18, "K": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# VAT Summary
# ---------------------------------------------------------------------------


def _build_vat_declaration_summary(
    *,
    invoices: list[Invoice],
    company: Company,
    year: int,
    period: int,
    period_type: str,
    previous_vat_credit: int,
    import_purchase_value: int,
    import_purchase_vat: int,
    deductible_input_vat_override: int | None,
    adjustment_decrease: int,
    adjustment_increase: int,
    transferred_vat_credit: int,
    investment_project_offset_vat: int,
    refund_requested_vat: int,
) -> dict:
    purchase_annex, sales_annex = _build_annexes(invoices, company)

    by_rate = {
        VATRate.ZERO.value: {"rate": VATRate.ZERO.value, "input_amount": 0, "output_amount": 0, "input_vat": 0, "output_vat": 0},
        VATRate.FIVE.value: {"rate": VATRate.FIVE.value, "input_amount": 0, "output_amount": 0, "input_vat": 0, "output_vat": 0},
        VATRate.EIGHT.value: {"rate": VATRate.EIGHT.value, "input_amount": 0, "output_amount": 0, "input_vat": 0, "output_vat": 0},
        VATRate.TEN.value: {"rate": VATRate.TEN.value, "input_amount": 0, "output_amount": 0, "input_vat": 0, "output_vat": 0},
        VATRate.EXEMPT.value: {"rate": VATRate.EXEMPT.value, "input_amount": 0, "output_amount": 0, "input_vat": 0, "output_vat": 0},
        VATRate.NOT_APPLICABLE.value: {"rate": VATRate.NOT_APPLICABLE.value, "input_amount": 0, "output_amount": 0, "input_vat": 0, "output_vat": 0},
    }

    sales = {
        "exempt_base": 0,
        "rate_0_base": 0,
        "rate_5_base": 0,
        "rate_5_vat": 0,
        "standard_rate_base": 0,
        "standard_rate_vat": 0,
        "not_taxable_base": 0,
    }
    purchases = {
        "purchase_base": 0,
        "purchase_vat": 0,
    }
    validation_issues: list[str] = []

    if not validate_mst(company.tax_code):
        validation_issues.append("Company MST failed checksum validation.")

    for invoice in invoices:
        direction = _invoice_direction(invoice, company)
        rate = invoice.vat_rate.value if hasattr(invoice.vat_rate, "value") else str(invoice.vat_rate)
        bucket = by_rate.setdefault(rate, {"rate": rate, "input_amount": 0, "output_amount": 0, "input_vat": 0, "output_vat": 0})
        subtotal_amount = int(invoice.subtotal_amount or 0)
        vat_amount = int(invoice.vat_amount or 0)

        if direction == "purchase":
            bucket["input_amount"] += subtotal_amount
            bucket["input_vat"] += vat_amount
            purchases["purchase_base"] += subtotal_amount
            purchases["purchase_vat"] += vat_amount
        else:
            bucket["output_amount"] += subtotal_amount
            bucket["output_vat"] += vat_amount
            if invoice.vat_rate == VATRate.EXEMPT:
                sales["exempt_base"] += subtotal_amount
            elif invoice.vat_rate == VATRate.ZERO:
                sales["rate_0_base"] += subtotal_amount
            elif invoice.vat_rate == VATRate.FIVE:
                sales["rate_5_base"] += subtotal_amount
                sales["rate_5_vat"] += vat_amount
            elif invoice.vat_rate in {VATRate.EIGHT, VATRate.TEN}:
                sales["standard_rate_base"] += subtotal_amount
                sales["standard_rate_vat"] += vat_amount
            else:
                sales["not_taxable_base"] += subtotal_amount

        for tax_code, label in (
            (invoice.seller_tax_code, "seller"),
            (invoice.buyer_tax_code, "buyer"),
        ):
            normalized = normalize_tax_code(tax_code)
            if normalized and not validate_mst(normalized):
                invoice_ref = invoice.invoice_number or invoice.id
                validation_issues.append(f"Invoice {invoice_ref} has invalid {label} MST.")

    field_23 = purchases["purchase_base"] + import_purchase_value
    field_24 = purchases["purchase_vat"] + import_purchase_vat
    field_25 = deductible_input_vat_override if deductible_input_vat_override is not None else field_24
    field_26 = sales["exempt_base"]
    field_29 = sales["rate_0_base"]
    field_30 = sales["rate_5_base"]
    field_31 = sales["rate_5_vat"]
    field_32 = sales["standard_rate_base"]
    field_33 = sales["standard_rate_vat"]
    field_32a = sales["not_taxable_base"]
    field_27 = field_29 + field_30 + field_32 + field_32a
    field_28 = field_31 + field_33
    field_34 = field_26 + field_27
    field_35 = field_28
    field_36 = field_35 - field_25
    basis_amount = field_36 - previous_vat_credit + adjustment_decrease - adjustment_increase - transferred_vat_credit
    field_40a = max(0, basis_amount)
    field_41 = max(0, -basis_amount)
    field_40b = min(investment_project_offset_vat, field_40a)
    field_40 = field_40a - field_40b
    field_42 = min(refund_requested_vat, field_41)
    field_43 = field_41 - field_42
    no_activity = (
        field_23 == 0 and field_24 == 0 and field_34 == 0 and previous_vat_credit == 0 and field_42 == 0 and field_40 == 0
    )

    filing_fields = {
        "21": no_activity,
        "22": previous_vat_credit,
        "23": field_23,
        "23a": import_purchase_value,
        "24": field_24,
        "24a": import_purchase_vat,
        "25": field_25,
        "26": field_26,
        "27": field_27,
        "28": field_28,
        "29": field_29,
        "30": field_30,
        "31": field_31,
        "32": field_32,
        "32a": field_32a,
        "33": field_33,
        "34": field_34,
        "35": field_35,
        "36": field_36,
        "37": adjustment_decrease,
        "38": adjustment_increase,
        "39a": transferred_vat_credit,
        "40a": field_40a,
        "40b": field_40b,
        "40": field_40,
        "41": field_41,
        "42": field_42,
        "43": field_43,
    }

    # Build vat_by_rate as dict keyed by float rate
    vat_by_rate_dict: dict[float, int] = {}
    for bucket in by_rate.values():
        rate_str = bucket["rate"]
        if rate_str in ("0", "5", "8", "10"):
            vat_by_rate_dict[float(rate_str) / 100] = bucket["output_vat"]

    # Build warnings list
    warnings: list[dict] = []
    for issue in (detect_missing_mst(invoices), detect_duplicate_invoices(invoices),
                  detect_low_confidence(invoices), detect_vat_mismatch(invoices)):
        if issue:
            warnings.append({"type": issue.type, "message": issue.message, "invoice_ids": issue.invoice_ids})

    return {
        "year": year,
        "period": period,
        "period_type": period_type,
        "input_vat_total": field_25,
        "output_vat_total": field_35,
        "net_vat": field_40 if field_40 > 0 else -field_43,
        "vat_by_rate": vat_by_rate_dict,
        "invoice_count": len(invoices),
        "filing_deadline": get_vat_declaration_deadline(year, period, period_type),
        "warnings": warnings,
        "generated_at": datetime.utcnow().isoformat(),
        "company_name": company.name,
        "tax_code": company.tax_code,
        "period_label": _period_label(period_type, period, year),
        # Legacy fields still returned for backward compat
        "payable_vat": field_40,
        "carry_forward_vat": field_43,
        "refund_requested_vat": field_42,
        "declaration_deadline": get_vat_declaration_deadline(year, period, period_type),
        "by_rate": list(by_rate.values()),
        "filing_fields": filing_fields,
        "inputs": {
            "previous_vat_credit": previous_vat_credit,
            "import_purchase_value": import_purchase_value,
            "import_purchase_vat": import_purchase_vat,
            "deductible_input_vat_override": deductible_input_vat_override,
            "adjustment_decrease": adjustment_decrease,
            "adjustment_increase": adjustment_increase,
            "transferred_vat_credit": transferred_vat_credit,
            "investment_project_offset_vat": investment_project_offset_vat,
            "refund_requested_vat": refund_requested_vat,
        },
        "purchase_annex": purchase_annex,
        "sales_annex": sales_annex,
        "validation_issues": sorted(set(validation_issues)),
    }


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.get("/vat-summary")
async def get_vat_summary(
    year: int = Query(..., ge=2020, le=2035),
    period: int = Query(..., ge=1, le=12, description="Month (1-12) or quarter (1-4)"),
    period_type: Literal["monthly", "quarterly"] = Query("quarterly"),
    previous_vat_credit: int = Query(0, ge=0, description="Filed [22] carried from the previous return"),
    import_purchase_value: int = Query(0, ge=0, description="Import purchase base included in [23a]"),
    import_purchase_vat: int = Query(0, ge=0, description="Import purchase VAT included in [24a]"),
    deductible_input_vat_override: int | None = Query(None, ge=0, description="Override for declaration field [25]"),
    adjustment_decrease: int = Query(0, ge=0, description="Declaration field [37]"),
    adjustment_increase: int = Query(0, ge=0, description="Declaration field [38]"),
    transferred_vat_credit: int = Query(0, ge=0, description="Declaration field [39a]"),
    investment_project_offset_vat: int = Query(0, ge=0, description="Declaration field [40b]"),
    refund_requested_vat: int = Query(0, ge=0, description="Declaration field [42]"),
    format: Literal["json", "excel"] = Query("json"),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    start_date, end_date = _period_range(year, period, period_type)
    invoices = await _period_invoices(db, current_user.company_id, start_date, end_date)
    company = await _company(db, current_user.company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    summary = _build_vat_declaration_summary(
        invoices=invoices,
        company=company,
        year=year,
        period=period,
        period_type=period_type,
        previous_vat_credit=previous_vat_credit,
        import_purchase_value=import_purchase_value,
        import_purchase_vat=import_purchase_vat,
        deductible_input_vat_override=deductible_input_vat_override,
        adjustment_decrease=adjustment_decrease,
        adjustment_increase=adjustment_increase,
        transferred_vat_credit=transferred_vat_credit,
        investment_project_offset_vat=investment_project_offset_vat,
        refund_requested_vat=refund_requested_vat,
    )

    if format == "excel":
        period_str = f"Q{period}" if period_type == "quarterly" else f"M{period:02d}"
        filename = f"vat-summary-{year}-{period_str}.xlsx"
        buffer = _build_vat_workbook(summary, company)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    return summary


@router.get("/sales-invoices")
async def get_sales_invoices(
    year: int = Query(..., ge=2020, le=2035),
    period: int = Query(..., ge=1, le=12),
    period_type: Literal["monthly", "quarterly"] = Query("quarterly"),
    format: Literal["json", "excel"] = Query("json"),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return all sales invoices (buyer = current company) for the period."""
    start_date, end_date = _period_range(year, period, period_type)
    invoices = await _period_invoices(db, current_user.company_id, start_date, end_date)
    company = await _company(db, current_user.company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    payload = _invoice_list_payload(invoices, company, period_type, period, year, direction_filter="sale")

    if format == "excel":
        period_str = f"Q{period}" if period_type == "quarterly" else f"M{period:02d}"
        filename = f"sales-invoices-{year}-{period_str}.xlsx"
        buffer = _build_invoice_list_workbook(payload["items"], "Sales Invoices", company.name)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    return payload


@router.get("/purchase-invoices")
async def get_purchase_invoices(
    year: int = Query(..., ge=2020, le=2035),
    period: int = Query(..., ge=1, le=12),
    period_type: Literal["monthly", "quarterly"] = Query("quarterly"),
    format: Literal["json", "excel"] = Query("json"),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return all purchase invoices (seller = current company) for the period."""
    start_date, end_date = _period_range(year, period, period_type)
    invoices = await _period_invoices(db, current_user.company_id, start_date, end_date)
    company = await _company(db, current_user.company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    payload = _invoice_list_payload(invoices, company, period_type, period, year, direction_filter="purchase")

    if format == "excel":
        period_str = f"Q{period}" if period_type == "quarterly" else f"M{period:02d}"
        filename = f"purchase-invoices-{year}-{period_str}.xlsx"
        buffer = _build_invoice_list_workbook(payload["items"], "Purchase Invoices", company.name)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    return payload


@router.get("/exceptions")
async def get_exceptions(
    year: int = Query(..., ge=2020, le=2035),
    period: int = Query(..., ge=1, le=12),
    period_type: Literal["monthly", "quarterly"] = Query("quarterly"),
    format: Literal["json", "excel"] = Query("json"),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Return data-quality issues across all invoices in the period.
    Covers: missing MST, duplicate invoice numbers, low extraction confidence, and VAT mismatches.
    """
    start_date, end_date = _period_range(year, period, period_type)
    invoices = await _period_invoices(db, current_user.company_id, start_date, end_date)
    company = await _company(db, current_user.company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    all_issues: list[dict] = []
    for detector in (
        detect_missing_mst,
        detect_duplicate_invoices,
        detect_low_confidence,
        detect_vat_mismatch,
    ):
        issue = detector(invoices)
        if issue:
            all_issues.append({
                "type": issue.type,
                "message": issue.message,
                "invoices": issue.invoice_ids,
                "count": issue.count,
            })

    result = {"issues": all_issues}

    if format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Exceptions"

        bold = Font(bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")

        ws.append(["Type", "Message", "Invoice IDs", "Count"])
        for cell in ws[1]:
            cell.font = bold
            cell.fill = header_fill
            cell.border = thin_border

        for issue in all_issues:
            ws.append([issue["type"], issue["message"], ", ".join(issue["invoices"]), issue["count"]])
            row_idx = ws.max_row
            for cell in ws[row_idx]:
                cell.border = thin_border

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 10

        period_str = f"Q{period}" if period_type == "quarterly" else f"M{period:02d}"
        filename = f"exceptions-{year}-{period_str}.xlsx"
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    return result


@router.get("/invoice-list")
async def get_invoice_list_report(
    year: int = Query(...),
    period: int = Query(...),
    period_type: str = Query("quarterly", pattern="^(monthly|quarterly)$"),
    invoice_direction: str = Query("all", pattern="^(purchase|sale|all)$"),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    start_date, end_date = _period_range(year, period, period_type)
    invoices = await _period_invoices(db, current_user.company_id, start_date, end_date)
    company = await _company(db, current_user.company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    purchase_annex, sales_annex = _build_annexes(invoices, company)
    if invoice_direction == "purchase":
        items = purchase_annex["items"]
    elif invoice_direction == "sale":
        items = sales_annex["items"]
    else:
        items = purchase_annex["items"] + sales_annex["items"]

    return {
        "year": year,
        "period": period,
        "period_type": period_type,
        "purchase_annex": purchase_annex,
        "sales_annex": sales_annex,
        "items": items,
        "total": len(items),
    }


@router.get("/cit-provisional")
async def get_cit_provisional(
    year: int = Query(...),
    quarter: int = Query(..., ge=1, le=4),
    non_deductible_expenses: int = Query(0, ge=0),
    loss_carried_forward: int = Query(0, ge=0),
    cit_paid_ytd: int = Query(0, ge=0),
    annual_cit_estimate: int | None = Query(None, ge=0),
    cit_rate: Decimal = Query(CIT_STANDARD_RATE, ge=0, le=1),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    company = await _company(db, current_user.company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    ytd_end = _quarter_end_date(year, quarter)
    journal_entries = await _ytd_posted_journal_entries(db, current_user.company_id, year, ytd_end)
    revenue, deductible_expenses, other_income, other_expenses = _aggregate_cit_bases(journal_entries)

    calculation = calculate_quarterly_cit_provision(
        revenue=revenue,
        deductible_expenses=deductible_expenses,
        other_income=other_income,
        other_expenses=other_expenses,
        non_deductible_expenses=non_deductible_expenses,
        loss_carried_forward=loss_carried_forward,
        cit_paid_ytd=cit_paid_ytd,
        cit_rate=Decimal(str(cit_rate)),
        annual_cit_estimate=annual_cit_estimate,
        quarter=quarter,
    )

    return {
        "year": year,
        "quarter": quarter,
        "revenue": calculation.revenue,
        "deductible_expenses": calculation.deductible_expenses,
        "other_income": calculation.other_income,
        "other_expenses": calculation.other_expenses,
        "accounting_profit": calculation.accounting_profit,
        "non_deductible_expenses": calculation.non_deductible_expenses,
        "loss_carried_forward": calculation.loss_carried_forward,
        "taxable_income": calculation.taxable_income,
        "cit_rate": float(calculation.cit_rate),
        "cit_amount": calculation.cit_liability_ytd,
        "already_paid": calculation.cit_paid_ytd,
        "amount_due": calculation.amount_due,
        "annual_cit_estimate": calculation.annual_cit_estimate,
        "minimum_cumulative_payment": calculation.minimum_cumulative_payment,
        "due_date": get_cit_quarter_payment_deadline(year, quarter),
    }


@router.get("/export/vat-declaration")
async def export_vat_declaration(
    year: int = Query(...),
    period: int = Query(...),
    period_type: str = Query("quarterly", pattern="^(monthly|quarterly)$"),
    format: str = Query("xlsx", pattern="^(xlsx|pdf)$"),
    previous_vat_credit: int = Query(0, ge=0),
    import_purchase_value: int = Query(0, ge=0),
    import_purchase_vat: int = Query(0, ge=0),
    deductible_input_vat_override: int | None = Query(None, ge=0),
    adjustment_decrease: int = Query(0, ge=0),
    adjustment_increase: int = Query(0, ge=0),
    transferred_vat_credit: int = Query(0, ge=0),
    investment_project_offset_vat: int = Query(0, ge=0),
    refund_requested_vat: int = Query(0, ge=0),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if format == "pdf":
        raise HTTPException(status_code=501, detail="PDF export is not implemented yet. Use xlsx.")

    start_date, end_date = _period_range(year, period, period_type)
    invoices = await _period_invoices(db, current_user.company_id, start_date, end_date)
    company = await _company(db, current_user.company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    summary = _build_vat_declaration_summary(
        invoices=invoices,
        company=company,
        year=year,
        period=period,
        period_type=period_type,
        previous_vat_credit=previous_vat_credit,
        import_purchase_value=import_purchase_value,
        import_purchase_vat=import_purchase_vat,
        deductible_input_vat_override=deductible_input_vat_override,
        adjustment_decrease=adjustment_decrease,
        adjustment_increase=adjustment_increase,
        transferred_vat_credit=transferred_vat_credit,
        investment_project_offset_vat=investment_project_offset_vat,
        refund_requested_vat=refund_requested_vat,
    )
    buffer = _build_vat_workbook(summary, company)
    filename = f"vat-declaration-{year}-{period_type}-{period}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
